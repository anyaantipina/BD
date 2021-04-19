from mrjob.job import MRJob 
from mrjob.step import MRStep
from mrjob.protocol import RawProtocol
import json
import os
import openpyxl
import xml.etree.ElementTree as ET
import re
from math import ceil

WORD_RE = re.compile(r"[\w']+")
SCRIPT_DIR = os.path.dirname(os.path.realpath(__file__))
INPUT_DATA_DIR = os.path.join(SCRIPT_DIR, "data")
DATA_SCHEMA_DIR = os.path.join(SCRIPT_DIR, "schema")
INTERNAL_DIR = os.path.join(SCRIPT_DIR, "private")
OUTPUT_DATA_DIR = os.path.join(SCRIPT_DIR, "out")
XLSX_FILENAME = "data-bars.xlsx"
XML_FILENAME = "data-coworkings.xml"
INTERNAL_FILE = "partition.txt"
TARGET_SCHEMA_FILE = "target-schema.json"

class Helper():

    def split_data(self):
        partition_file = open(os.path.join(INTERNAL_DIR, INTERNAL_FILE), "w")
        
        xlsx_source = openpyxl.load_workbook(os.path.join(INPUT_DATA_DIR, XLSX_FILENAME))
        sheet = xlsx_source.active
        self.split(partition_file, sheet.max_row, "xlsx", 2)
        partition_file.write("\n")
        
        xml_source = ET.parse(os.path.join(INPUT_DATA_DIR, XML_FILENAME))
        root = xml_source.getroot()
        self.split(partition_file, len(list(root)), "xml", 0)

        partition_file.close()
        
    def split(self, partition_file, count, format, start_pos):
        per_mapper = ceil(count / 10)
        i = start_pos
        while i < count:
            start = i
            end = i + per_mapper if (i + per_mapper <= count) else count
            partition_file.write("%s %s %s" % (start,end, format))
            i = end
            if i < count:
                partition_file.write("\n")

    def get_target_schema(self):
        json_file = open(os.path.join(DATA_SCHEMA_DIR, TARGET_SCHEMA_FILE), "r", encoding="utf-8")
        schema = json.load(json_file)
        return schema["items"]["properties"]

    def get_schema_converter(self):
        schema_converter = {}
        schema_converter["PublicPhone"] = "ContactPhone"
        schema_converter["ShortName"] = "Name"
        schema_converter["Latitude_WGS84"] = "GeoData"
        schema_converter["Longitude_WGS84"] = "GeoData"
        schema_converter["geoData"] = "GeoData"
        return schema_converter


class ConvertToJSON(MRJob):
    xlsx_source = openpyxl.load_workbook(os.path.join(INPUT_DATA_DIR, XLSX_FILENAME))
    xlsx_json_file = os.path.join(OUTPUT_DATA_DIR, "".join(XLSX_FILENAME.split(".")[:-1]) + ".json")
    xlsx_count = xlsx_source.active.max_row
    
    xml_source = ET.parse(os.path.join(INPUT_DATA_DIR, XML_FILENAME))
    xml_json_file = os.path.join(OUTPUT_DATA_DIR, "".join(XML_FILENAME.split(".")[:-1]) + ".json")
    xml_count = len(list(xml_source.getroot()))
    
    helper = Helper()
    schema_converter = helper.get_schema_converter()
    target_schema = helper.get_target_schema()
    
    def start(self):
        self.helper.split_data()
        self.start_impl(self.xlsx_json_file)
        self.start_impl(self.xml_json_file)

    def end(self):
        self.end_impl(self.xlsx_json_file, self.xlsx_count)
        self.end_impl(self.xml_json_file, self.xml_count)
            
    def start_impl(self, file):
        with open(file, "w", encoding="utf-8") as json_file:
            json_file.write("{\n\"Заведения\": [\n")
            
    def end_impl(self, file, count):
        with open(file, "a", encoding="utf-8") as json_file:
            json_file.write("\t{\n\t\t\"count\": " + "\"" + str(count) + "\"\n\t}\n]\n}")
    
    def get_target_schema_field(self, name):
        if not name in self.schema_converter.keys():
            return name
        return self.schema_converter[name]
        
    def mapper(self, _, line):
        sheet = self.xlsx_source.active
        range_rows = []
        for word in WORD_RE.findall(line):
            range_rows.append(word)
        range_rows[0] = int(range_rows[0])
        range_rows[1] = int(range_rows[1])
        primary_key = "Name" if range_rows[2] == "xlsx" else "ShortName"
        
        for elem in self.generate_mapper(range_rows):
            yield (range_rows[2], elem[primary_key]), elem
            
    def generate_mapper(self, range):
        if range[2] == "xlsx":
            return self.xlsx_mapper(range)
        return self.xml_mapper(range)
    
    def xlsx_mapper(self, row_range):
        sheet = self.xlsx_source.active
        elems = []
        for row in sheet.iter_rows(row_range[0], row_range[1]):
            elem = {}
            for column in range(1, sheet.max_column + 1):
                elem[sheet.cell(row=1,column=column).value] = row[column-1].value
            elems.append(elem)
        return elems
    
    def xml_mapper(self, range):
        elems = []
        for child in list(self.xml_source.getroot())[range[0]: range[1] + 1]:
            elem = {}
            for childchild in list(child):
                if len(list(childchild)) == 0:
                    elem[childchild.tag] = childchild.text
                else:
                    attribs = {}
                    for attrib in list(childchild):
                        if not attrib.tag in attribs.keys():
                            attribs[attrib.tag] = [attrib.text]
                        else:
                            attribs[attrib.tag].append(attrib.text)
                    elem[childchild.tag] = attribs
            elems.append(elem)
        return elems

    def reducer(self, format_key, values):
        file = self.xlsx_json_file if format_key[0] == "xlsx" else self.xml_json_file
        name = "Bar" if format_key[0] == "xlsx" else "Coworking"
        with open(file, "a", encoding="utf-8") as json_file:
            for value in values:
                json_file.write("\t{\n")
                json_file.write("\t\t\"type\": \"" + name + "\",\n")
                json_file.write("\t\t\"properties\": \n\t\t{\n")
                keys = list(value.keys())
                target_keys = []
                for key in keys:
                    if self.get_target_schema_field(key) in self.target_schema.keys():
                        target_keys.append(key)
                
                self.dump_geoData(json_file, target_keys, value, format_key[0])
                for key in target_keys[:-1]:
                    self.dump(json_file, key, value[key], "\",\n")
                self.dump(json_file, target_keys[-1], value[target_keys[-1]], "\"\n")
                json_file.write("\t\t}\n\t},\n")
                return
            
    def dump(self, json_file, key, value, last):
        if isinstance(value, dict):
            temp_keys = list(value.keys())
            if self.target_schema[key]["type"] != "DICTIONARY" or len(value[temp_keys[0]]) == 1:
                str_value = ""
                if isinstance(value[temp_keys[0]], list):
                    str_value = value[temp_keys[0]][0]
                else:
                    str_value = value[temp_keys[0]]
                json_file.write("\t\t\t\"" + self.get_target_schema_field(key) + "\": \"" + str_value.replace('"', "'") + last)
            else:
                json_file.write("\t\t\t\"" + self.get_target_schema_field(key) + "\": \n\t\t\t{\n")
                for k in temp_keys[:-1]:
                    json_file.write("\t\t\t\t\"" + self.get_target_schema_field(k) + "\": \"" + value[k].replace('"', "'") +"\",\n")
                if len(temp_keys) > 0:
                    if isinstance(value[temp_keys[-1]], list):
                        json_file.write("\t\t\t\t\"" + self.get_target_schema_field(temp_keys[-1]) + "\": \"" + ",".join(value[temp_keys[-1]]).replace('"', "'") +"\"\n")
                    else:
                        json_file.write("\t\t\t\t\"" + self.get_target_schema_field(temp_keys[-1]) + "\": \"" + value[temp_keys[-1]].replace('"', "'") +"\"\n")
                json_file.write("\t\t\t}\n\t\t},\n")
        else:
            json_file.write("\t\t\t\"" + self.get_target_schema_field(key) + "\": \"" + value.replace('"', "'") + last)
            
    def dump_geoData(self, json_file, target_keys, values, format):
        json_file.write("\t\t\t\"GeoData\": \n\t\t\t{\n")
        
        if (format == "xlsx"):
            if "Latitude_WGS84" in target_keys:
                json_file.write("\t\t\t\t\"Latitude_WGS84\": \"" + values["Latitude_WGS84"] + "\",\n")
                target_keys.remove("Latitude_WGS84")
                    
            if "Longitude_WGS84" in target_keys:
                json_file.write("\t\t\t\t\"Longitude_WGS84\": \"" + values["Longitude_WGS84"] + "\"\n")
                target_keys.remove("Longitude_WGS84")

        if (format == "xml") and "geoData" in target_keys:
            json_file.write("\t\t\t\t\"Latitude_WGS84\": \"" + values["geoData"]["coordinates"][1] +"\",\n")
            json_file.write("\t\t\t\t\"Longitude_WGS84\": \"" + values["geoData"]["coordinates"][0] + "\"\n")
        
        target_keys.remove("geoData")
        json_file.write("\t\t\t},\n")
            
    
    def steps(self):
        return [
            MRStep(mapper=self.mapper,
                   reducer=self.reducer)
        ]

if __name__ == '__main__':

    converter = ConvertToJSON()
    converter.start()
    converter.run()
    converter.end()
