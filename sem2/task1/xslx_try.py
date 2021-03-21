
import os
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.realpath(__file__))
INPUT_DATA_DIR = os.path.join(SCRIPT_DIR, "data")
XLSX_FILENAME = "data-bars.xlsx"


xlsx_source_file = openpyxl.load_workbook(os.path.join(INPUT_DATA_DIR, XLSX_FILENAME))
sheet = xlsx_source_file.active
print(sheet.max_row, sheet.max_column)
for column in sheet.iter_cols(1, sheet.max_column):
    print(column[0].value)

xlsx_source_file.close()
